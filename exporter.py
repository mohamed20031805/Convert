import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
import os

def exporter_excel(output_df, entete, chemin_pdf):
    if output_df is None or output_df.empty:
        print("⚠️  Aucune donnée à exporter !")
        return

    nom_pdf     = os.path.splitext(os.path.basename(chemin_pdf))[0]
    date_today  = datetime.today().strftime("%Y%m%d")
    nom_fichier = f"output/{nom_pdf}_{date_today}.xlsx"

    with pd.ExcelWriter(nom_fichier, engine="openpyxl") as writer:
        output_df.to_excel(writer, index=False, sheet_name="Open Positions", startrow=5)
        ws = writer.sheets["Open Positions"]

        label_font = Font(bold=True, size=11)
        ws["A1"] = "Broker"           ; ws["B1"] = entete.get("Broker", "")
        ws["A2"] = "Client"           ; ws["B2"] = entete.get("Client", "")
        ws["A3"] = "Close of Business"; ws["B3"] = entete.get("Close of Business", "")
        ws["A4"] = "Account"          ; ws["B4"] = entete.get("Account", "")

        for cell in ["A1","A2","A3","A4"]:
            ws[cell].font = label_font

        formatter(ws, "1F4E79", start_row=6)

    print(f"\n✅ Fichier Excel créé : {nom_fichier}")
    print(f"📊 {len(output_df)} lignes exportées")
    return nom_fichier


def formatter(worksheet, couleur_header, start_row=1):
    header_fill = PatternFill("solid", fgColor=couleur_header)
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[start_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for col in worksheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        worksheet.column_dimensions[col[0].column_letter].width = max_len + 4